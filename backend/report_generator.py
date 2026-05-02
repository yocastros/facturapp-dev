import os
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.formatting.rule import DataBarRule
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False

# Paleta corporativa
COLOR_AZUL_OSCURO = "1B3A5C"
COLOR_DORADO = "C9A84C"
COLOR_GRIS_CLARO = "F5F3EE"
COLOR_BLANCO = "FFFFFF"
COLOR_GRIS_MED = "D9D5CC"
COLOR_VERDE = "2E7D32"
COLOR_ROJO = "C62828"
COLOR_AZUL_CLARO = "E8EEF4"


def _borde_fino():
    lado = Side(style='thin', color="CCCCCC")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _borde_medio():
    lado = Side(style='medium', color=COLOR_AZUL_OSCURO)
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _celda_cabecera(ws, fila, col, texto, ancho=None):
    celda = ws.cell(row=fila, column=col, value=texto)
    celda.font = Font(bold=True, color=COLOR_BLANCO, name='Calibri', size=11)
    celda.fill = PatternFill("solid", fgColor=COLOR_AZUL_OSCURO)
    celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    celda.border = _borde_fino()
    return celda


def _celda_dato(ws, fila, col, valor, formato=None, color_fondo=None, negrita=False, alineacion='left'):
    celda = ws.cell(row=fila, column=col, value=valor)
    celda.font = Font(name='Calibri', size=10, bold=negrita,
                      color=COLOR_AZUL_OSCURO)
    celda.alignment = Alignment(horizontal=alineacion, vertical='center')
    celda.border = _borde_fino()
    if color_fondo:
        celda.fill = PatternFill("solid", fgColor=color_fondo)
    if formato:
        celda.number_format = formato
    return celda


def generar_reporte_excel(documentos, ruta_salida, fecha_desde=None, fecha_hasta=None):
    """Genera reporte Excel completo con múltiples hojas."""
    if not EXCEL_DISPONIBLE:
        return None, "openpyxl no está instalado"

    wb = openpyxl.Workbook()

    # Filtrar por fechas si se especifican
    docs_filtrados = documentos
    # (filtrado ya viene del controlador)

    facturas = [d for d in docs_filtrados if d['tipo'] == 'factura']
    albaranes = [d for d in docs_filtrados if d['tipo'] == 'albaran']

    total_facturas = sum(d['total'] or 0 for d in facturas)
    total_albaranes = sum(d['total'] or 0 for d in albaranes)
    total_global = total_facturas + total_albaranes

    # ─── HOJA 1: PORTADA ────────────────────────────────────────────────────
    ws_portada = wb.active
    ws_portada.title = "Portada"
    ws_portada.sheet_view.showGridLines = False

    # Cabecera principal
    ws_portada.merge_cells('A1:H3')
    cab = ws_portada['A1']
    cab.value = "SISTEMA DE GESTIÓN DE FACTURAS Y ALBARANES"
    cab.font = Font(bold=True, color=COLOR_BLANCO, name='Calibri', size=18)
    cab.fill = PatternFill("solid", fgColor=COLOR_AZUL_OSCURO)
    cab.alignment = Alignment(horizontal='center', vertical='center')

    ws_portada.merge_cells('A4:H4')
    sub = ws_portada['A4']
    sub.value = f"Reporte generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    sub.font = Font(color=COLOR_AZUL_OSCURO, name='Calibri', size=11, italic=True)
    sub.fill = PatternFill("solid", fgColor=COLOR_DORADO)
    sub.alignment = Alignment(horizontal='center', vertical='center')

    if fecha_desde or fecha_hasta:
        ws_portada.merge_cells('A5:H5')
        filtro = ws_portada['A5']
        filtro.value = f"Período: {fecha_desde or '—'} a {fecha_hasta or '—'}"
        filtro.font = Font(color=COLOR_AZUL_OSCURO, name='Calibri', size=10)
        filtro.fill = PatternFill("solid", fgColor=COLOR_GRIS_CLARO)
        filtro.alignment = Alignment(horizontal='center', vertical='center')

    # KPIs
    kpis = [
        ("Total Documentos", len(docs_filtrados), None),
        ("Facturas", len(facturas), None),
        ("Albaranes", len(albaranes), None),
        ("Importe Facturas", total_facturas, '#,##0.00 €'),
        ("Importe Albaranes", total_albaranes, '#,##0.00 €'),
        ("TOTAL GLOBAL", total_global, '#,##0.00 €'),
    ]

    fila_kpi = 7
    ws_portada.merge_cells(f'A{fila_kpi}:H{fila_kpi}')
    titulo_kpi = ws_portada[f'A{fila_kpi}']
    titulo_kpi.value = "MÉTRICAS CLAVE"
    titulo_kpi.font = Font(bold=True, color=COLOR_BLANCO, name='Calibri', size=12)
    titulo_kpi.fill = PatternFill("solid", fgColor=COLOR_AZUL_OSCURO)
    titulo_kpi.alignment = Alignment(horizontal='center', vertical='center')

    for i, (label, valor, fmt) in enumerate(kpis):
        fila = fila_kpi + 1 + (i // 3) * 3
        col_base = 1 + (i % 3) * 3

        col_fin = col_base + 1
        ws_portada.merge_cells(
            start_row=fila, start_column=col_base,
            end_row=fila, end_column=col_fin
        )
        lbl_cell = ws_portada.cell(row=fila, column=col_base, value=label)
        lbl_cell.font = Font(bold=True, color=COLOR_BLANCO, name='Calibri', size=10)
        lbl_cell.fill = PatternFill("solid", fgColor=COLOR_AZUL_OSCURO)
        lbl_cell.alignment = Alignment(horizontal='center', vertical='center')
        lbl_cell.border = _borde_fino()

        ws_portada.merge_cells(
            start_row=fila + 1, start_column=col_base,
            end_row=fila + 1, end_column=col_fin
        )
        val_cell = ws_portada.cell(row=fila + 1, column=col_base, value=valor)
        val_cell.font = Font(bold=True, color=COLOR_AZUL_OSCURO, name='Calibri', size=14)
        val_cell.fill = PatternFill("solid", fgColor=COLOR_GRIS_CLARO)
        val_cell.alignment = Alignment(horizontal='center', vertical='center')
        val_cell.border = _borde_medio()
        if fmt:
            val_cell.number_format = fmt

    # Ajustar columnas portada
    for col in range(1, 9):
        ws_portada.column_dimensions[get_column_letter(col)].width = 16
    for fila in [1, 2, 3]:
        ws_portada.row_dimensions[fila].height = 20
    ws_portada.row_dimensions[4].height = 18

    # ─── HOJA 2: LISTADO COMPLETO ────────────────────────────────────────────
    ws_lista = wb.create_sheet("Listado Completo")
    ws_lista.sheet_view.showGridLines = False

    # Cabecera
    ws_lista.merge_cells('A1:J1')
    cab2 = ws_lista['A1']
    cab2.value = "LISTADO COMPLETO DE DOCUMENTOS"
    cab2.font = Font(bold=True, color=COLOR_BLANCO, name='Calibri', size=14)
    cab2.fill = PatternFill("solid", fgColor=COLOR_AZUL_OSCURO)
    cab2.alignment = Alignment(horizontal='center', vertical='center')
    ws_lista.row_dimensions[1].height = 30

    cabeceras = ['ID', 'Tipo', 'Número', 'Fecha', 'Proveedor', 'CIF',
                 'Base Imponible', 'IVA', 'Total', 'Estado']
    for col, cab in enumerate(cabeceras, 1):
        _celda_cabecera(ws_lista, 2, col, cab)
    ws_lista.row_dimensions[2].height = 22

    for fila_idx, doc in enumerate(docs_filtrados, 3):
        color = COLOR_GRIS_CLARO if fila_idx % 2 == 0 else COLOR_BLANCO
        _celda_dato(ws_lista, fila_idx, 1, doc['id'], color_fondo=color, alineacion='center')
        _celda_dato(ws_lista, fila_idx, 2,
                    'FACTURA' if doc['tipo'] == 'factura' else 'ALBARÁN',
                    color_fondo=color, alineacion='center')
        _celda_dato(ws_lista, fila_idx, 3, doc.get('numero') or '', color_fondo=color)
        _celda_dato(ws_lista, fila_idx, 4, doc.get('fecha') or '', color_fondo=color)
        _celda_dato(ws_lista, fila_idx, 5, doc.get('proveedor') or '', color_fondo=color)
        _celda_dato(ws_lista, fila_idx, 6, doc.get('cif') or '', color_fondo=color)
        _celda_dato(ws_lista, fila_idx, 7, doc.get('base_imponible') or 0,
                    formato='#,##0.00 €', color_fondo=color, alineacion='right')
        _celda_dato(ws_lista, fila_idx, 8, doc.get('iva') or 0,
                    formato='#,##0.00 €', color_fondo=color, alineacion='right')
        _celda_dato(ws_lista, fila_idx, 9, doc.get('total') or 0,
                    formato='#,##0.00 €', color_fondo=color, negrita=True, alineacion='right')

        estado = doc.get('estado', '')
        col_estado = ws_lista.cell(row=fila_idx, column=10, value=estado)
        col_estado.font = Font(name='Calibri', size=10, bold=True,
                               color=COLOR_VERDE if estado == 'PROCESADO'
                               else (COLOR_DORADO if estado == 'FACTURA_ASOCIADA' else COLOR_ROJO))
        col_estado.alignment = Alignment(horizontal='center', vertical='center')
        col_estado.fill = PatternFill("solid", fgColor=color)
        col_estado.border = _borde_fino()

    # Totales
    ultima_fila = len(docs_filtrados) + 3
    _celda_dato(ws_lista, ultima_fila, 1, 'TOTALES', negrita=True,
                color_fondo=COLOR_DORADO, alineacion='center')
    for c in range(2, 7):
        ws_lista.cell(row=ultima_fila, column=c).fill = PatternFill("solid", fgColor=COLOR_DORADO)
        ws_lista.cell(row=ultima_fila, column=c).border = _borde_fino()

    sum_base = sum(d.get('base_imponible') or 0 for d in docs_filtrados)
    sum_iva = sum(d.get('iva') or 0 for d in docs_filtrados)
    sum_total = sum(d.get('total') or 0 for d in docs_filtrados)
    _celda_dato(ws_lista, ultima_fila, 7, sum_base,
                formato='#,##0.00 €', negrita=True, color_fondo=COLOR_DORADO, alineacion='right')
    _celda_dato(ws_lista, ultima_fila, 8, sum_iva,
                formato='#,##0.00 €', negrita=True, color_fondo=COLOR_DORADO, alineacion='right')
    _celda_dato(ws_lista, ultima_fila, 9, sum_total,
                formato='#,##0.00 €', negrita=True, color_fondo=COLOR_DORADO, alineacion='right')

    # Anchos columnas
    anchos = [6, 12, 18, 14, 30, 14, 16, 14, 16, 18]
    for i, ancho in enumerate(anchos, 1):
        ws_lista.column_dimensions[get_column_letter(i)].width = ancho

    # ─── HOJA 3: RESUMEN ─────────────────────────────────────────────────────
    ws_res = wb.create_sheet("Resumen")
    ws_res.sheet_view.showGridLines = False

    ws_res.merge_cells('A1:F1')
    r1 = ws_res['A1']
    r1.value = "RESUMEN POR TIPO DE DOCUMENTO"
    r1.font = Font(bold=True, color=COLOR_BLANCO, name='Calibri', size=14)
    r1.fill = PatternFill("solid", fgColor=COLOR_AZUL_OSCURO)
    r1.alignment = Alignment(horizontal='center', vertical='center')
    ws_res.row_dimensions[1].height = 30

    cabeceras_res = ['Tipo', 'Cantidad', 'Base Imponible', 'IVA Total', 'Total', '% del Total']
    for col, c in enumerate(cabeceras_res, 1):
        _celda_cabecera(ws_res, 2, col, c)

    filas_res = [
        ('FACTURAS', len(facturas),
         sum(d.get('base_imponible') or 0 for d in facturas),
         sum(d.get('iva') or 0 for d in facturas),
         total_facturas),
        ('ALBARANES', len(albaranes),
         sum(d.get('base_imponible') or 0 for d in albaranes),
         sum(d.get('iva') or 0 for d in albaranes),
         total_albaranes),
        ('TOTAL', len(docs_filtrados),
         sum(d.get('base_imponible') or 0 for d in docs_filtrados),
         sum(d.get('iva') or 0 for d in docs_filtrados),
         total_global),
    ]
    colores_res = [COLOR_AZUL_CLARO, COLOR_GRIS_CLARO, COLOR_DORADO]

    for fila_idx, (tipo, cant, base, iva_val, tot) in enumerate(filas_res, 3):
        color = colores_res[fila_idx - 3]
        negrita = fila_idx == 5
        _celda_dato(ws_res, fila_idx, 1, tipo, color_fondo=color, negrita=negrita)
        _celda_dato(ws_res, fila_idx, 2, cant, color_fondo=color, negrita=negrita, alineacion='center')
        _celda_dato(ws_res, fila_idx, 3, base, formato='#,##0.00 €',
                    color_fondo=color, negrita=negrita, alineacion='right')
        _celda_dato(ws_res, fila_idx, 4, iva_val, formato='#,##0.00 €',
                    color_fondo=color, negrita=negrita, alineacion='right')
        _celda_dato(ws_res, fila_idx, 5, tot, formato='#,##0.00 €',
                    color_fondo=color, negrita=negrita, alineacion='right')
        pct = round((tot / total_global * 100) if total_global > 0 else 0, 1)
        _celda_dato(ws_res, fila_idx, 6, f"{pct}%", color_fondo=color,
                    negrita=negrita, alineacion='center')

    for i, ancho in enumerate([18, 12, 18, 14, 16, 14], 1):
        ws_res.column_dimensions[get_column_letter(i)].width = ancho

    # ─── HOJA 4: NETEO ──────────────────────────────────────────────────────
    ws_neteo = wb.create_sheet("Neteo Facturas-Albaranes")
    ws_neteo.sheet_view.showGridLines = False

    ws_neteo.merge_cells('A1:G1')
    n1 = ws_neteo['A1']
    n1.value = "NETEO FACTURAS ↔ ALBARANES"
    n1.font = Font(bold=True, color=COLOR_BLANCO, name='Calibri', size=14)
    n1.fill = PatternFill("solid", fgColor=COLOR_AZUL_OSCURO)
    n1.alignment = Alignment(horizontal='center', vertical='center')
    ws_neteo.row_dimensions[1].height = 30

    cab_neteo = ['Factura Nº', 'Proveedor', 'Fecha Factura',
                 'Albarán Asociado', 'Fecha Albarán', 'Total', 'Estado Neteo']
    for col, c in enumerate(cab_neteo, 1):
        _celda_cabecera(ws_neteo, 2, col, c)

    fila_n = 3
    for doc in docs_filtrados:
        if doc['tipo'] == 'factura':
            albaranes_asoc = doc.get('albaranes_asociados', [])
            if albaranes_asoc:
                for alb in albaranes_asoc:
                    color = COLOR_AZUL_CLARO
                    _celda_dato(ws_neteo, fila_n, 1, doc.get('numero') or '', color_fondo=color)
                    _celda_dato(ws_neteo, fila_n, 2, doc.get('proveedor') or '', color_fondo=color)
                    _celda_dato(ws_neteo, fila_n, 3, doc.get('fecha') or '', color_fondo=color)
                    _celda_dato(ws_neteo, fila_n, 4, alb.get('numero') or '', color_fondo=color)
                    _celda_dato(ws_neteo, fila_n, 5, alb.get('fecha') or '', color_fondo=color)
                    _celda_dato(ws_neteo, fila_n, 6, doc.get('total') or 0,
                                formato='#,##0.00 €', color_fondo=color, alineacion='right')
                    estado_n = ws_neteo.cell(row=fila_n, column=7, value='✓ NETEADO')
                    estado_n.font = Font(bold=True, color=COLOR_VERDE, name='Calibri', size=10)
                    estado_n.fill = PatternFill("solid", fgColor=color)
                    estado_n.alignment = Alignment(horizontal='center', vertical='center')
                    estado_n.border = _borde_fino()
                    fila_n += 1
            else:
                color = COLOR_GRIS_CLARO
                _celda_dato(ws_neteo, fila_n, 1, doc.get('numero') or '', color_fondo=color)
                _celda_dato(ws_neteo, fila_n, 2, doc.get('proveedor') or '', color_fondo=color)
                _celda_dato(ws_neteo, fila_n, 3, doc.get('fecha') or '', color_fondo=color)
                _celda_dato(ws_neteo, fila_n, 4, '—', color_fondo=color, alineacion='center')
                _celda_dato(ws_neteo, fila_n, 5, '—', color_fondo=color, alineacion='center')
                _celda_dato(ws_neteo, fila_n, 6, doc.get('total') or 0,
                            formato='#,##0.00 €', color_fondo=color, alineacion='right')
                est = ws_neteo.cell(row=fila_n, column=7, value='⚠ PENDIENTE')
                est.font = Font(bold=True, color=COLOR_ROJO, name='Calibri', size=10)
                est.fill = PatternFill("solid", fgColor=color)
                est.alignment = Alignment(horizontal='center', vertical='center')
                est.border = _borde_fino()
                fila_n += 1

    for i, ancho in enumerate([18, 28, 14, 18, 14, 14, 16], 1):
        ws_neteo.column_dimensions[get_column_letter(i)].width = ancho

    # Guardar
    os.makedirs(os.path.dirname(ruta_salida), exist_ok=True)
    wb.save(ruta_salida)
    return ruta_salida, None


def generar_reporte_contable(docs, ruta_salida, nombre_proveedor='Todos los proveedores'):
    """Excel contable de facturas: hoja Resumen por proveedor + hoja Detalle."""
    if not EXCEL_DISPONIBLE:
        return None, "openpyxl no está instalado"

    from collections import defaultdict

    wb = openpyxl.Workbook()

    # ─── HOJA 1: RESUMEN ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:E1')
    t = ws['A1']
    t.value = "INFORME CONTABLE DE FACTURAS"
    t.font = Font(bold=True, color=COLOR_AZUL_OSCURO, name='Calibri', size=16)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:E2')
    ws['A2'].value = f"Proveedor: {nombre_proveedor}"
    ws['A2'].font = Font(name='Calibri', size=11, color=COLOR_AZUL_OSCURO)
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('A3:E3')
    ws['A3'].value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A3'].font = Font(name='Calibri', size=10, color=COLOR_AZUL_OSCURO, italic=True)
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center')

    for col, cab in enumerate(['Proveedor', 'Nº Facturas', 'Base Imponible', 'IVA (€)', 'Total'], 1):
        _celda_cabecera(ws, 5, col, cab)
    ws.row_dimensions[5].height = 22

    grupos = defaultdict(lambda: {'count': 0, 'base': 0.0, 'iva': 0.0, 'total': 0.0})
    for doc in docs:
        k = doc.get('proveedor') or 'Sin proveedor'
        grupos[k]['count'] += 1
        grupos[k]['base']  += doc.get('base_imponible') or 0
        grupos[k]['iva']   += doc.get('iva') or 0
        grupos[k]['total'] += doc.get('total') or 0

    fila = 6
    for i, (prov, vals) in enumerate(sorted(grupos.items())):
        color = COLOR_BLANCO if i % 2 == 0 else 'F5F5F5'
        _celda_dato(ws, fila, 1, prov, color_fondo=color)
        _celda_dato(ws, fila, 2, vals['count'], color_fondo=color, alineacion='center')
        _celda_dato(ws, fila, 3, vals['base'],  formato='€ #,##0.00', color_fondo=color, alineacion='right')
        _celda_dato(ws, fila, 4, vals['iva'],   formato='€ #,##0.00', color_fondo=color, alineacion='right')
        _celda_dato(ws, fila, 5, vals['total'], formato='€ #,##0.00', color_fondo=color, alineacion='right')
        fila += 1

    _celda_dato(ws, fila, 1, 'TOTALES', negrita=True, color_fondo=COLOR_DORADO, alineacion='center')
    _celda_dato(ws, fila, 2, sum(v['count'] for v in grupos.values()), negrita=True, color_fondo=COLOR_DORADO, alineacion='center')
    _celda_dato(ws, fila, 3, sum(v['base']  for v in grupos.values()), formato='€ #,##0.00', negrita=True, color_fondo=COLOR_DORADO, alineacion='right')
    _celda_dato(ws, fila, 4, sum(v['iva']   for v in grupos.values()), formato='€ #,##0.00', negrita=True, color_fondo=COLOR_DORADO, alineacion='right')
    _celda_dato(ws, fila, 5, sum(v['total'] for v in grupos.values()), formato='€ #,##0.00', negrita=True, color_fondo=COLOR_DORADO, alineacion='right')

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 12
    for col_ltr in ['C', 'D', 'E']:
        ws.column_dimensions[col_ltr].width = 18

    # ─── HOJA 2: DETALLE FACTURAS ────────────────────────────────────────────
    ws2 = wb.create_sheet("Detalle Facturas")
    ws2.sheet_view.showGridLines = False

    cabs2 = ['Nº Factura', 'Fecha', 'Proveedor', 'CIF',
             'Base Imponible', 'IVA %', 'IVA €', 'Total', 'Neteo']
    for col, cab in enumerate(cabs2, 1):
        _celda_cabecera(ws2, 1, col, cab)
    ws2.row_dimensions[1].height = 22

    docs_sorted = sorted(docs, key=lambda d: (d.get('proveedor') or '', d.get('fecha') or ''))
    for i, doc in enumerate(docs_sorted, 2):
        color = COLOR_BLANCO if i % 2 == 0 else COLOR_GRIS_CLARO
        _celda_dato(ws2, i, 1, doc.get('numero') or '—', color_fondo=color)
        _celda_dato(ws2, i, 2, doc.get('fecha') or '—', color_fondo=color)
        _celda_dato(ws2, i, 3, doc.get('proveedor') or '—', color_fondo=color)
        _celda_dato(ws2, i, 4, doc.get('cif') or '—', color_fondo=color)
        _celda_dato(ws2, i, 5, doc.get('base_imponible') or 0, formato='€ #,##0.00', color_fondo=color, alineacion='right')
        pct = (doc.get('porcentaje_iva') or 0) / 100
        _celda_dato(ws2, i, 6, pct, formato='0.0%', color_fondo=color, alineacion='center')
        _celda_dato(ws2, i, 7, doc.get('iva') or 0, formato='€ #,##0.00', color_fondo=color, alineacion='right')
        _celda_dato(ws2, i, 8, doc.get('total') or 0, formato='€ #,##0.00', color_fondo=color, negrita=True, alineacion='right')
        neteo_val = '✓ Con albarán' if doc.get('estado') == 'FACTURA_ASOCIADA' else '✗ Sin albarán'
        cn = _celda_dato(ws2, i, 9, neteo_val, color_fondo=color, alineacion='center')
        cn.font = Font(name='Calibri', size=10, bold=True,
                       color=COLOR_VERDE if doc.get('estado') == 'FACTURA_ASOCIADA' else COLOR_ROJO)

    ultima = len(docs_sorted) + 2
    for c in range(1, 10):
        ws2.cell(row=ultima, column=c).fill = PatternFill("solid", fgColor=COLOR_DORADO)
        ws2.cell(row=ultima, column=c).border = _borde_fino()
    _celda_dato(ws2, ultima, 1, 'TOTALES', negrita=True, color_fondo=COLOR_DORADO)
    _celda_dato(ws2, ultima, 5, sum(d.get('base_imponible') or 0 for d in docs_sorted),
                formato='€ #,##0.00', negrita=True, color_fondo=COLOR_DORADO, alineacion='right')
    _celda_dato(ws2, ultima, 7, sum(d.get('iva') or 0 for d in docs_sorted),
                formato='€ #,##0.00', negrita=True, color_fondo=COLOR_DORADO, alineacion='right')
    _celda_dato(ws2, ultima, 8, sum(d.get('total') or 0 for d in docs_sorted),
                formato='€ #,##0.00', negrita=True, color_fondo=COLOR_DORADO, alineacion='right')

    for i, ancho in enumerate([18, 12, 35, 14, 16, 8, 14, 16, 16], 1):
        ws2.column_dimensions[get_column_letter(i)].width = ancho

    os.makedirs(os.path.dirname(ruta_salida), exist_ok=True)
    wb.save(ruta_salida)
    return ruta_salida, None


def generar_reporte_analitico(docs_con_lineas, ruta_salida, nombre_proveedor='Todos los proveedores'):
    """Excel analítico: Portada + Detalle por línea + Coste Promedio Ponderado."""
    if not EXCEL_DISPONIBLE:
        return None, "openpyxl no está instalado"

    import re
    from collections import defaultdict, Counter

    def _norm(s):
        return re.sub(r'\s+', ' ', (s or '').lower().strip())

    wb = openpyxl.Workbook()

    # ─── HOJA 1: PORTADA ─────────────────────────────────────────────────────
    ws_p = wb.active
    ws_p.title = "Portada"
    ws_p.sheet_view.showGridLines = False

    ws_p.merge_cells('B2:H3')
    t = ws_p['B2']
    t.value = "ANÁLISIS DE COMPRAS POR PRODUCTO"
    t.font = Font(bold=True, color=COLOR_AZUL_OSCURO, name='Calibri', size=20)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws_p.row_dimensions[2].height = 30
    ws_p.row_dimensions[3].height = 30

    ws_p.merge_cells('B4:H4')
    ws_p['B4'].value = "Coste Promedio Ponderado"
    ws_p['B4'].font = Font(name='Calibri', size=12, color=COLOR_DORADO)
    ws_p['B4'].alignment = Alignment(horizontal='center', vertical='center')

    ws_p.merge_cells('B5:H5')
    ws_p['B5'].value = f"Proveedor: {nombre_proveedor}"
    ws_p['B5'].font = Font(name='Calibri', size=11, color=COLOR_AZUL_OSCURO)
    ws_p['B5'].alignment = Alignment(horizontal='center', vertical='center')

    ws_p.merge_cells('B6:H6')
    ws_p['B6'].value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_p['B6'].font = Font(name='Calibri', size=10, color=COLOR_AZUL_OSCURO, italic=True)
    ws_p['B6'].alignment = Alignment(horizontal='center', vertical='center')

    total_docs    = len(docs_con_lineas)
    total_lineas  = sum(len(d.get('lineas', [])) for d in docs_con_lineas)
    importe_total = sum(d.get('total') or 0 for d in docs_con_lineas)
    num_provs     = len(set(d.get('proveedor') or '' for d in docs_con_lineas))

    kpis_p = [
        ('Total Documentos', total_docs, None),
        ('Total Líneas', total_lineas, None),
        ('Importe Total', importe_total, '#,##0.00 €'),
        ('Nº Proveedores', num_provs, None),
    ]
    kpi_cols_p = [2, 4, 6, 8]
    for i, (label, valor, fmt_num) in enumerate(kpis_p):
        col = kpi_cols_p[i]
        lbl = ws_p.cell(row=8, column=col, value=label)
        lbl.font = Font(bold=True, color=COLOR_BLANCO, name='Calibri', size=10)
        lbl.fill = PatternFill("solid", fgColor=COLOR_AZUL_OSCURO)
        lbl.alignment = Alignment(horizontal='center', vertical='center')
        lbl.border = _borde_medio()
        val = ws_p.cell(row=9, column=col, value=valor)
        val.font = Font(bold=True, color=COLOR_AZUL_OSCURO, name='Calibri', size=16)
        val.fill = PatternFill("solid", fgColor=COLOR_GRIS_CLARO)
        val.alignment = Alignment(horizontal='center', vertical='center')
        val.border = _borde_medio()
        if fmt_num:
            val.number_format = fmt_num

    ws_p.row_dimensions[8].height = 20
    ws_p.row_dimensions[9].height = 30
    for col in range(1, 10):
        ws_p.column_dimensions[get_column_letter(col)].width = 18

    # ─── HOJA 2: DETALLE POR DOCUMENTO ───────────────────────────────────────
    ws_d = wb.create_sheet("Detalle por Documento")
    ws_d.sheet_view.showGridLines = False

    cabs_d = ['Proveedor', 'Nº Documento', 'Fecha', 'Descripción Producto',
              'Cantidad', 'Unidad', 'Precio Unitario', 'Importe Línea']
    for col, cab in enumerate(cabs_d, 1):
        _celda_cabecera(ws_d, 1, col, cab)
    ws_d.row_dimensions[1].height = 22

    docs_sorted_d = sorted(docs_con_lineas, key=lambda d: (
        d.get('proveedor') or '', d.get('fecha') or ''
    ))

    fila_d = 2
    for doc_idx, doc in enumerate(docs_sorted_d):
        lineas = doc.get('lineas', [])
        if not lineas:
            continue
        color = COLOR_BLANCO if doc_idx % 2 == 0 else COLOR_GRIS_CLARO
        for linea_idx, linea in enumerate(lineas):
            prov_v  = (doc.get('proveedor') or '—') if linea_idx == 0 else ''
            num_v   = (doc.get('numero') or f"#{doc.get('id','')}") if linea_idx == 0 else ''
            fecha_v = (doc.get('fecha') or '—') if linea_idx == 0 else ''
            _celda_dato(ws_d, fila_d, 1, prov_v,  color_fondo=color)
            _celda_dato(ws_d, fila_d, 2, num_v,   color_fondo=color)
            _celda_dato(ws_d, fila_d, 3, fecha_v, color_fondo=color)
            _celda_dato(ws_d, fila_d, 4, linea.get('descripcion') or '', color_fondo=color)
            _celda_dato(ws_d, fila_d, 5, linea.get('cantidad') or 0,
                        formato='#,##0.000', color_fondo=color, alineacion='right')
            _celda_dato(ws_d, fila_d, 6, linea.get('unidad') or '',
                        color_fondo=color, alineacion='center')
            _celda_dato(ws_d, fila_d, 7, linea.get('precio_unitario') or 0,
                        formato='€ #,##0.00', color_fondo=color, alineacion='right')
            _celda_dato(ws_d, fila_d, 8, linea.get('importe_linea') or 0,
                        formato='€ #,##0.00', color_fondo=color, alineacion='right')
            fila_d += 1

    for i, ancho in enumerate([30, 18, 12, 40, 14, 10, 16, 16], 1):
        ws_d.column_dimensions[get_column_letter(i)].width = ancho

    # ─── HOJA 3: COSTE PROMEDIO PONDERADO ────────────────────────────────────
    ws_cpp = wb.create_sheet("Coste Promedio Ponderado")
    ws_cpp.sheet_view.showGridLines = False

    cabs_cpp = ['Producto', 'Total Unidades', 'Importe Total', 'CPP (€/ud)', 'Nº Compras', 'Proveedores']
    for col, cab in enumerate(cabs_cpp, 1):
        _celda_cabecera(ws_cpp, 1, col, cab)
    ws_cpp.row_dimensions[1].height = 22

    grupos_cpp = defaultdict(lambda: {
        'descripciones': [], 'cantidad': 0.0,
        'importe': 0.0, 'count': 0, 'proveedores': set()
    })
    for doc in docs_con_lineas:
        prov = doc.get('proveedor') or ''
        for linea in doc.get('lineas', []):
            key = _norm(linea.get('descripcion', ''))
            if not key:
                continue
            g = grupos_cpp[key]
            g['descripciones'].append(linea.get('descripcion', ''))
            g['cantidad'] += linea.get('cantidad') or 0
            g['importe']  += linea.get('importe_linea') or 0
            g['count']    += 1
            if prov:
                g['proveedores'].add(prov)

    rows_cpp = []
    for key, g in grupos_cpp.items():
        most_freq = Counter(g['descripciones']).most_common(1)[0][0]
        cant = g['cantidad']
        imp  = g['importe']
        rows_cpp.append({
            'producto': most_freq,
            'cantidad': cant,
            'importe': imp,
            'cpp': imp / cant if cant > 0 else 0,
            'count': g['count'],
            'proveedores': ', '.join(sorted(g['proveedores'])),
        })
    rows_cpp.sort(key=lambda x: x['importe'], reverse=True)

    for i, row in enumerate(rows_cpp, 2):
        is_top10 = (i - 2) < 10
        color = 'FFF9C4' if is_top10 else (COLOR_BLANCO if i % 2 == 0 else COLOR_GRIS_CLARO)
        _celda_dato(ws_cpp, i, 1, row['producto'],   color_fondo=color)
        _celda_dato(ws_cpp, i, 2, row['cantidad'],   formato='#,##0.000',    color_fondo=color, alineacion='right')
        _celda_dato(ws_cpp, i, 3, row['importe'],    formato='€ #,##0.00',   color_fondo=color, alineacion='right')
        _celda_dato(ws_cpp, i, 4, row['cpp'],        formato='€ #,##0.0000', color_fondo=color, alineacion='right')
        _celda_dato(ws_cpp, i, 5, row['count'],      color_fondo=color, alineacion='center')
        _celda_dato(ws_cpp, i, 6, row['proveedores'], color_fondo=color)

    for i, ancho in enumerate([45, 16, 16, 16, 12, 40], 1):
        ws_cpp.column_dimensions[get_column_letter(i)].width = ancho

    os.makedirs(os.path.dirname(ruta_salida), exist_ok=True)
    wb.save(ruta_salida)
    return ruta_salida, None
